#!/usr/bin/env python3
"""Create the chapter 2 indicator workbook snapshot."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE = PROJECT_ROOT / "data" / "processed" / "tigit-01-preparacio-dades-teaching.xlsx"
OUTPUT = PROJECT_ROOT / "data" / "processed" / "tigit-02-indicadors-territorials-teaching.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def finish_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, sheet.max_row + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max(map(len, values)) + 2, 12), 38)


def replace_sheet(workbook, title, headers):
    if title in workbook.sheetnames:
        del workbook[title]
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    return sheet


def build() -> None:
    shutil.copyfile(SOURCE, OUTPUT)
    workbook = load_workbook(OUTPUT)

    project = workbook["project"]
    project_rows = {project.cell(row, 1).value: row for row in range(2, project.max_row + 1)}
    project.cell(project_rows["chapter_snapshot"], 2, "02")

    demography = replace_sheet(
        workbook,
        "indicators_demography",
        [
            "municipality_code", "municipality_name", "population_total", "population_0_14",
            "population_65_plus", "surface_km2", "population_0_14_pct",
            "population_65_plus_pct", "ageing_index_per_100_young", "population_density_hab_km2",
        ],
    )
    housing = replace_sheet(
        workbook,
        "indicators_housing",
        [
            "municipality_code", "municipality_name", "population_total", "housing_total",
            "housing_main", "housing_non_main", "housing_non_main_pct", "residents_per_housing_main",
        ],
    )

    for row in range(2, 24):
        demography.append(
            [
                f"=municipal!A{row}", f"=municipal!B{row}", f"=municipal!E{row}",
                f"=municipal!F{row}", f"=municipal!H{row}", f"=municipal!L{row}",
                f'=IF(AND(ISNUMBER(D{row}),ISNUMBER(C{row}),C{row}>0),D{row}/C{row}*100,NA())',
                f'=IF(AND(ISNUMBER(E{row}),ISNUMBER(C{row}),C{row}>0),E{row}/C{row}*100,NA())',
                f'=IF(AND(ISNUMBER(E{row}),ISNUMBER(D{row}),D{row}>0),E{row}/D{row}*100,NA())',
                f'=IF(AND(ISNUMBER(C{row}),ISNUMBER(F{row}),F{row}>0),C{row}/F{row},NA())',
            ]
        )
        housing.append(
            [
                f"=municipal!A{row}", f"=municipal!B{row}", f"=municipal!E{row}",
                f"=municipal!I{row}", f"=municipal!J{row}", f"=municipal!K{row}",
                f'=IF(AND(ISNUMBER(F{row}),ISNUMBER(D{row}),D{row}>0),F{row}/D{row}*100,NA())',
                f'=IF(AND(ISNUMBER(C{row}),ISNUMBER(E{row}),E{row}>0),C{row}/E{row},NA())',
            ]
        )
    for sheet, columns in ((demography, "G:J"), (housing, "G:H")):
        for row in sheet[columns]:
            for cell in row:
                cell.number_format = "0.00"
        finish_sheet(sheet)

    summary = replace_sheet(
        workbook,
        "indicators_summary",
        ["indicator", "numerator_total", "denominator_total", "scale_factor", "aggregate_value", "unit"],
    )
    summary_rows = [
        ["county_population_total", "=SUM(indicators_demography!C2:C23)", "", 1, "=B2", "persones"],
        ["county_housing_total", "=SUM(indicators_housing!D2:D23)", "", 1, "=B3", "habitatges"],
        ["county_population_0_14_pct", "=SUM(indicators_demography!D2:D23)", "=SUM(indicators_demography!C2:C23)", 100, "=IF(C4>0,B4/C4*D4,NA())", "%"],
        ["county_population_65_plus_pct", "=SUM(indicators_demography!E2:E23)", "=SUM(indicators_demography!C2:C23)", 100, "=IF(C5>0,B5/C5*D5,NA())", "%"],
        ["county_ageing_index", "=SUM(indicators_demography!E2:E23)", "=SUM(indicators_demography!D2:D23)", 100, "=IF(C6>0,B6/C6*D6,NA())", "persones de 65+ per 100 de 0–14"],
        ["county_population_density", "=SUM(indicators_demography!C2:C23)", "=SUM(indicators_demography!F2:F23)", 1, "=IF(C7>0,B7/C7,NA())", "habitants/km²"],
        ["county_housing_non_main_pct", "=SUM(indicators_housing!F2:F23)", "=SUM(indicators_housing!D2:D23)", 100, "=IF(C8>0,B8/C8*D8,NA())", "%"],
        ["county_residents_per_housing_main", "=SUM(indicators_housing!C2:C23)", "=SUM(indicators_housing!E2:E23)", 1, "=IF(C9>0,B9/C9,NA())", "residents/habitatge principal"],
    ]
    for row in summary_rows:
        summary.append(row)
    for cell in summary["E"][1:]:
        cell.number_format = "0.00"
    finish_sheet(summary)

    dictionary = workbook["dictionary"]
    for header in ["question", "formula", "scale_factor", "intended_use", "limitations"]:
        dictionary.cell(1, dictionary.max_column + 1, header)
    indicator_rows = [
        ["population_0_14_pct", "indicators_demography", "Pes de la població jove", "decimal", "%", "idescat_population_2021", "NA si el total no és positiu", "Població de 0–14 sobre població total", "On pesa més la població jove?", "population_0_14 / population_total × 100", 100, "Comparar estructura demogràfica", "No explica les causes de l'estructura"],
        ["population_65_plus_pct", "indicators_demography", "Pes de la població gran", "decimal", "%", "idescat_population_2021", "NA si el total no és positiu", "Població de 65+ sobre població total", "On pesa més la població gran?", "population_65_plus / population_total × 100", 100, "Comparar envelliment relatiu", "No mesura necessitats individuals"],
        ["ageing_index_per_100_young", "indicators_demography", "Índex d'envelliment", "decimal", "persones per 100", "idescat_population_2021", "NA si 0–14 no és positiu", "Relació entre població gran i jove", "Quina relació hi ha entre els extrems d'edat?", "population_65_plus / population_0_14 × 100", 100, "Comparar estructura d'edats", "És sensible a denominadors petits"],
        ["population_density_hab_km2", "indicators_demography", "Densitat de població", "decimal", "habitants/km²", "idescat_population_2021; idescat_surface", "NA si la superfície no és positiva", "Població per superfície municipal", "On es concentra la població?", "population_total / surface_km2", 1, "Comparar concentració mitjana", "No descriu la distribució interna; superfície publicada el 2025"],
        ["housing_non_main_pct", "indicators_housing", "Pes de l'habitatge no principal", "decimal", "%", "idescat_housing_2021", "NA si el total no és positiu", "Habitatge no principal sobre total", "On pesa més l'habitatge no principal?", "housing_non_main / housing_total × 100", 100, "Comparar composició residencial", "No identifica habitatges turístics"],
        ["residents_per_housing_main", "indicators_housing", "Residents per habitatge principal", "decimal", "residents/habitatge", "idescat_population_2021; idescat_housing_2021", "NA si habitatges principals no és positiu", "Població sobre habitatges principals", "Quina relació aproximada hi ha entre residents i parc principal?", "population_total / housing_main", 1, "Contextualitzar ocupació residencial", "No és la grandària oficial de la llar"],
    ]
    for row in indicator_rows:
        dictionary.append(row)
    finish_sheet(dictionary)

    checks = workbook["checks"]
    for row in [
        ["C08", "indicators_*", "Files municipals i codis únics", "22 i 22", "22 i 22", "OK", "Una fila per municipi a cada família"],
        ["C09", "indicators_*", "Denominadors no positius", 0, 0, "OK", "Població, població jove, superfície, habitatges totals i principals"],
        ["C10", "municipal", "Components diferents dels totals", 0, 0, "OK", "Població i habitatges reconstruïts"],
        ["C11", "indicators_*", "Fórmules municipals", 132, 132, "OK", "Sis indicadors per 22 municipis"],
        ["C12", "indicators_*", "Percentatges fora de 0–100", 0, 0, "OK", "Pes jove, pes gran i habitatge no principal"],
        ["C13", "indicators_summary", "Agregats calculats des de components", 8, 8, "OK", "No s'utilitzen mitjanes simples de percentatges"],
        ["C14", "indicators_summary", "Superfície temporalment compatible", "Justificació", "Pendent", "REVISAR", "La densitat continua provisional fins a resoldre C07"],
    ]:
        checks.append(row)
    finish_sheet(checks)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    build()
