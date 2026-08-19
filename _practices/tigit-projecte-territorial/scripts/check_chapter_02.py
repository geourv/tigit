#!/usr/bin/env python3
"""Check the chapter 2 indicator workbook."""

from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = PROJECT_ROOT / "data" / "processed" / "tigit-02-indicadors-territorials-teaching.xlsx"


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=False)
    assert {"indicators_demography", "indicators_housing", "indicators_summary"}.issubset(workbook.sheetnames)
    assert workbook["indicators_demography"].max_row == 23
    assert workbook["indicators_housing"].max_row == 23
    assert workbook["indicators_summary"].max_row == 9
    formulas = 0
    for sheet_name, columns in (("indicators_demography", range(7, 11)), ("indicators_housing", range(7, 9))):
        sheet = workbook[sheet_name]
        formulas += sum(str(sheet.cell(row, column).value).startswith("=") for row in range(2, 24) for column in columns)
    assert formulas == 132
    assert all(str(workbook["indicators_summary"].cell(row, 5).value).startswith("=") for row in range(2, 10))
    assert workbook["project"]["B3"].value == "02"
    print("Chapter 2 checks passed")


if __name__ == "__main__":
    main()
