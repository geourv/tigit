#!/usr/bin/env python3
"""Recalculate the workbook in Calc and export its QGIS transfer table."""

from __future__ import annotations

import csv
import subprocess
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REPOSITORY_ROOT = PROJECT_ROOT.parents[1]
SOURCE = PROJECT_ROOT / "data" / "processed" / "tigit-02-indicadors-territorials-teaching.xlsx"
TMP_DIR = REPOSITORY_ROOT / "tmp" / "qgis-transfer"
RECALCULATED = TMP_DIR / SOURCE.name
CSV_OUTPUT = PROJECT_ROOT / "data" / "processed" / "municipal-indicators-tarragones-2021.csv"
CSVT_OUTPUT = CSV_OUTPUT.with_suffix(".csvt")


def main() -> None:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    if RECALCULATED.exists():
        RECALCULATED.unlink()
    subprocess.run([
        "libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", str(TMP_DIR), str(SOURCE)
    ], check=True)

    workbook = load_workbook(RECALCULATED, data_only=True)
    municipal = workbook["municipal"]
    demography = workbook["indicators_demography"]
    housing = workbook["indicators_housing"]
    headers = [
        "mun_code", "municipality", "county_code", "year", "population_total",
        "population_65_plus", "population_65_plus_pct", "housing_total",
        "housing_non_main", "housing_non_main_pct", "indicator_status",
    ]
    rows = []
    for row in range(2, 24):
        values = [
            municipal.cell(row, 1).value, municipal.cell(row, 2).value, municipal.cell(row, 3).value,
            2021, municipal.cell(row, 5).value, municipal.cell(row, 8).value,
            demography.cell(row, 8).value, municipal.cell(row, 9).value,
            municipal.cell(row, 11).value, housing.cell(row, 7).value,
        ]
        status = "ok" if all(value is not None for value in values[4:]) else "missing_component"
        rows.append(values + [status])
    if len(rows) != 22 or len({row[0] for row in rows}) != 22:
        raise ValueError("QGIS transfer table must contain 22 unique municipalities")
    if not any(row[0] == "431711" and row[1] == "Vila-seca" for row in rows):
        raise ValueError("Vila-seca control row is missing")

    with CSV_OUTPUT.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)
    CSVT_OUTPUT.write_text(
        '"String","String","String","Integer","Integer","Integer","Real","Integer","Integer","Real","String"\n',
        encoding="ascii",
    )
    print(CSV_OUTPUT)


if __name__ == "__main__":
    main()
