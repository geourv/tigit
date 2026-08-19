#!/usr/bin/env python3
"""Check the chapter 7 palette workbook and vector proofs."""

from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = PROJECT_ROOT / "data" / "processed" / "tigit-07-teoria-color-teaching.xlsx"
FIGURES = PROJECT_ROOT / "dist" / "teaching" / "chapter-07" / "figures"


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=False)
    assert "palette" in workbook.sheetnames
    assert workbook["project"]["B3"].value == "07"
    assert workbook["palette"].max_row >= 26
    assert str(workbook["palette"]["F2"].value).startswith("=IF(")
    assert str(workbook["palette"]["L2"].value).startswith("=IF(")
    assert len(workbook["pivot_county_control"]._pivots) == 1
    assert len(workbook["pivot_population_age_sex"]._pivots) == 1
    for stem in (
        "palette-proof-age-structure-tarragones-2021",
        "palette-proof-non-principal-housing-tarragones-2021",
        "palette-proof-population-pyramid-vila-seca-2021",
    ):
        assert (FIGURES / f"{stem}.pdf").stat().st_size > 10_000
        ElementTree.parse(FIGURES / f"{stem}.svg")
    print("Chapter 7 checks passed")


if __name__ == "__main__":
    main()
