#!/usr/bin/env python3
"""Create the chapter 1 cumulative workbook scaffold."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "processed" / "tigit-01-preparacio-dades.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")


def add_sheet(workbook, title, headers, rows=()):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 2)}"

    for column, header in enumerate(headers, start=1):
        values = [str(header)] + [str(sheet.cell(row, column).value or "") for row in range(2, sheet.max_row + 1)]
        width = min(max(max(map(len, values)) + 2, 12), 42)
        sheet.column_dimensions[get_column_letter(column)].width = width

    return sheet


def build_workbook(output: Path, force: bool) -> None:
    if output.exists() and not force:
        raise SystemExit(f"Refusing to overwrite existing workbook: {output}")

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    project = add_sheet(
        workbook,
        "project",
        ["property", "value", "notes"],
        [
            ["project_name", "tigit-projecte-territorial", "Identitat estable del projecte"],
            ["chapter_snapshot", "01", "Fita de preparació de dades"],
            ["case_territory", "Tarragonès", "Cas de demostració; contrastar l'ortografia amb la font oficial"],
            ["territorial_unit", "municipi", "Una fila analítica per municipi"],
            ["reference_period", "2021", "Període representat per les dades de població i habitatge"],
            ["working_language", "ca", "El català és la llengua de treball"],
            ["spreadsheet_application", "", "Anotar Calc o Excel i la versió utilitzada"],
            ["regional_configuration", "", "Anotar els separadors decimals i de fórmules"],
            ["author", "", "Completar abans de produir captures docents"],
            ["question", "Com es distribueixen la població i el parc d'habitatges entre els municipis?", "Pregunta descriptiva inicial"],
        ],
    )
    project.column_dimensions["B"].width = 68
    project.column_dimensions["C"].width = 52

    source_headers = [
        "source_id",
        "producer",
        "title",
        "url",
        "reference_period",
        "access_date",
        "license",
        "format",
        "selection",
        "raw_file",
        "notes",
    ]
    add_sheet(
        workbook,
        "sources",
        source_headers,
        [
            [
                "idescat_population_2021",
                "Idescat",
                "Població. Per sexe i edat any a any",
                "https://www.idescat.cat/pub/?id=censph&n=10&lang=ca",
                "2021",
                "",
                "",
                "",
                "Municipis del Tarragonès; 2021; edats simples; total de tots els sexes",
                "",
                "La versió docent conserva també la consulta reproduïble de l'API",
            ],
            [
                "idescat_housing_2021",
                "Idescat",
                "Habitatges. Per tipus d'habitatge",
                "https://www.idescat.cat/pub/?id=censph&n=30&lang=ca",
                "2021",
                "",
                "",
                "",
                "Tots els municipis; total; principals; no principals",
                "",
                "Revisar les notes de confidencialitat i arrodoniment",
            ],
            [
                "idescat_codes",
                "Idescat",
                "Codis de municipis i comarques",
                "https://www.idescat.cat/codis/?id=50&n=9&lang=ca",
                "",
                "",
                "",
                "",
                "Municipis i comarca corresponent",
                "",
                "Els codis municipals s'han d'importar com a text",
            ],
            [
                "idescat_surface",
                "Idescat",
                "Superfície, densitat i entitats singulars",
                "https://www.idescat.cat/pub/?id=inddt&n=396&lang=ca",
                "",
                "",
                "",
                "",
                "Superfície municipal",
                "",
                "Registrar el període o la data d'actualització de la taula",
            ],
        ],
    )

    dictionary_rows = [
        ["municipality_code", "municipal", "Codi de municipi", "text", "", "idescat_codes", "No substituir per zero", "Clau oficial de sis caràcters per a les unions"],
        ["municipality_name", "municipal", "Municipi", "text", "", "idescat_codes", "No substituir", "Nom oficial del municipi"],
        ["county_code", "municipal", "Codi de comarca", "text", "", "idescat_codes", "No substituir", "Clau utilitzada per filtrar el territori d'estudi"],
        ["county_name", "municipal", "Comarca", "text", "", "idescat_codes", "No substituir", "Nom oficial de la comarca"],
        ["population_total", "municipal", "Població total", "integer", "persones", "idescat_population_2021", "Preservar absent", "Població resident durant el període de referència"],
        ["population_0_14", "municipal", "Població de 0 a 14 anys", "integer", "persones", "idescat_population_2021", "Preservar absent", "Població d'edats joves"],
        ["population_15_64", "municipal", "Població de 15 a 64 anys", "integer", "persones", "idescat_population_2021", "Preservar absent", "Interval d'edat central de la demostració"],
        ["population_65_plus", "municipal", "Població de 65 anys o més", "integer", "persones", "idescat_population_2021", "Preservar absent", "Població d'edats avançades"],
        ["housing_total", "municipal", "Habitatges totals", "integer", "habitatges", "idescat_housing_2021", "Preservar absent", "Parc total d'habitatges"],
        ["housing_main", "municipal", "Habitatges principals", "integer", "habitatges", "idescat_housing_2021", "Preservar absent", "Habitatges principals"],
        ["housing_non_main", "municipal", "Habitatges no principals", "integer", "habitatges", "idescat_housing_2021", "Preservar absent", "Habitatges d'ús ocasional i buits; no equivalen a habitatges turístics"],
        ["surface_km2", "municipal", "Superfície", "decimal", "km²", "idescat_surface", "Preservar absent", "Superfície municipal que després actuarà com a denominador"],
    ]
    add_sheet(
        workbook,
        "dictionary",
        ["field", "sheet", "label", "type", "unit", "source_id", "missing_treatment", "definition"],
        dictionary_rows,
    )

    add_sheet(
        workbook,
        "source_codes",
        ["Codi", "Nom", "Codi comarca", "Nom comarca"],
    )
    add_sheet(
        workbook,
        "source_surface",
        ["Codi", "Nom", "Superfície (km²)", "Densitat (hab./km²)", "Entitats singulars població (nombre)"],
    )
    add_sheet(
        workbook,
        "source_population",
        ["YEAR", "MUN", "AGE", "SEX", "CONCEPT", "value", "status"],
    )
    add_sheet(
        workbook,
        "source_housing",
        ["Codi", "Nom", "Habitatges familiars convencionals principals", "Habitatges familiars convencionals no principals", "Habitatges familiars convencionals", "Habitatges familiars no convencionals (allotjaments)", "Total habitatges familiars", "Establiments col·lectius"],
    )
    add_sheet(
        workbook,
        "prepared_codes",
        ["raw_municipality", "municipality_code", "municipality_name", "county_code", "county_name", "import_notes"],
    )
    add_sheet(
        workbook,
        "prepared_surface",
        ["raw_municipality", "municipality_code", "municipality_name", "surface_km2", "value_status", "import_notes"],
    )
    add_sheet(
        workbook,
        "prepared_population",
        [
            "year",
            "municipality_code",
            "municipality_name",
            "age_code",
            "age_label",
            "age_years",
            "population_total_sex",
            "value_status",
            "import_notes",
        ],
    )
    add_sheet(
        workbook,
        "prepared_housing",
        [
            "raw_municipality",
            "municipality_code",
            "municipality_name",
            "housing_total",
            "housing_main",
            "housing_non_main",
            "value_status",
            "import_notes",
        ],
    )

    checks = add_sheet(
        workbook,
        "checks",
        ["check_id", "scope", "test", "expected", "observed", "status", "notes"],
        [
            ["C01", "source_codes", "Files de dades importades", "> 0", "", "PENDENT", "Completar després de la importació"],
            ["C02", "prepared_codes", "Codis municipals únics", "Un codi per municipi", "", "PENDENT", "Utilitzar un filtre o una taula dinàmica"],
            ["C03", "municipal", "Files de la comarca seleccionada", "Contrastar amb la font oficial", "", "PENDENT", "Filtrar pel codi de comarca, no només pel nom"],
            ["C04", "municipal", "Codis municipals duplicats", "0", "", "PENDENT", "No eliminar duplicats sense diagnosticar-los"],
            ["C05", "municipal", "Valors obligatoris absents", "0 o documentats", "", "PENDENT", "Distingir una dada no disponible d'un zero observat"],
            ["C06", "municipal", "Grups d'edat respecte del total publicat", "Explicar les diferències", "", "PENDENT", "No forçar la coincidència de components arrodonits"],
        ],
    )
    for cell in checks[2]:
        cell.fill = SECTION_FILL

    add_sheet(
        workbook,
        "municipal",
        [
            "municipality_code",
            "municipality_name",
            "county_code",
            "county_name",
            "population_total",
            "population_0_14",
            "population_15_64",
            "population_65_plus",
            "housing_total",
            "housing_main",
            "housing_non_main",
            "surface_km2",
            "preparation_notes",
        ],
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(output)


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--force", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    build_workbook(args.output, args.force)
    print(args.output)
