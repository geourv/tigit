#!/usr/bin/env python3
"""Check chapter 3 workbook and vector exports."""

from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = PROJECT_ROOT / "data" / "processed" / "tigit-03-semiologia-visualitzacio-teaching.xlsx"
FIGURES = PROJECT_ROOT / "dist" / "teaching" / "chapter-03" / "figures"


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=False)
    expected = {
        "charts_data", "chart_00_audit", "chart_01_age_structure",
        "chart_02_nonprincipal", "chart_03_scatter", "chart_04_housing_donut",
        "chart_05_population_time", "source_population_time", "prepared_population_time",
        "chart_06_population_pyramid", "source_population_pyramid", "prepared_population_pyramid",
        "pivot_population_age_sex", "pyramid_data", "chart_07_housing_histogram", "histogram_data",
    }
    assert expected.issubset(workbook.sheetnames)
    assert workbook["project"]["B3"].value == "03"
    assert workbook["charts_data"].max_row == 23
    assert len(workbook["chart_00_audit"]._charts) == 2
    assert len(workbook["chart_01_age_structure"]._charts) == 1
    assert len(workbook["chart_02_nonprincipal"]._charts) == 1
    assert len(workbook["chart_03_scatter"]._charts) == 1
    assert len(workbook["chart_04_housing_donut"]._charts) == 1
    assert len(workbook["chart_05_population_time"]._charts) == 1
    assert len(workbook["chart_06_population_pyramid"]._charts) == 1
    assert len(workbook["chart_07_housing_histogram"]._charts) == 1
    assert len(workbook["pivot_county_control"]._pivots) == 1
    assert len(workbook["pivot_population_age_sex"]._pivots) == 1
    assert workbook["prepared_population_time"].max_row == 24
    assert workbook["pyramid_data"].max_row == 19
    scatter_series = workbook["chart_03_scatter"]._charts[0].series[0]
    assert scatter_series.graphicalProperties.line.noFill is True

    for name in (
        "age-structure-tarragones-2021.pdf",
        "non-principal-housing-tarragones-2021.pdf",
        "ageing-vs-non-principal-housing-tarragones-2021.pdf",
        "housing-composition-tarragones-2021.pdf",
        "population-vila-seca-2000-2022.pdf",
        "population-pyramid-vila-seca-2021.pdf",
        "housing-non-main-histogram-tarragones-2021.pdf",
    ):
        path = FIGURES / name
        assert path.stat().st_size > 10_000
    for name in (
        "age-structure-tarragones-2021.svg",
        "non-principal-housing-tarragones-2021.svg",
        "population-vila-seca-2000-2022.svg",
        "population-pyramid-vila-seca-2021.svg",
    ):
        ElementTree.parse(FIGURES / name)
    print("Chapter 3 checks passed")


if __name__ == "__main__":
    main()
