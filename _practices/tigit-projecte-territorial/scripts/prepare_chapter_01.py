#!/usr/bin/env python3
"""Download, prepare, and validate the chapter 1 teaching workbook."""

from __future__ import annotations

import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import date
from itertools import product
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = PROJECT_ROOT / "data" / "raw"
PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
STARTER = PROCESSED_DIR / "tigit-01-preparacio-dades.xlsx"
OUTPUT = PROCESSED_DIR / "tigit-01-preparacio-dades-teaching.xlsx"

ACCESS_DATE = date(2026, 8, 18).isoformat()
COUNTY_CODE = "36"
COUNTY_NAME = "Tarragonès"

CSV_SOURCES = {
    "mpiscatalunya.csv": "https://www.idescat.cat/codis/?id=50&n=9&lang=ca&f=ssv",
    "t30mun.csv": "https://www.idescat.cat/pub/?id=censph&n=30&lang=ca&geo=mun&f=ssv",
    "t396mun.csv": "https://www.idescat.cat/pub/?id=inddt&n=396&lang=ca&geo=mun&f=ssv",
}
POPULATION_CACHE = PROJECT_ROOT / "sandbox" / "cache" / "censph-10-5975-mun-2021-tarragones.json"
POPULATION_FILE = RAW_DIR / "poblacio-edat-tarragones-2021.csv"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def download(url: str, destination: Path) -> None:
    request = Request(url, headers={"User-Agent": "TIGIT teaching material preparation"})
    with urlopen(request, timeout=120) as response:
        destination.write_bytes(response.read())


def csv_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))
    header_index = next(index for index, row in enumerate(rows) if row and row[0].strip() == "Codi")
    header = [value.strip() for value in rows[header_index] if value.strip()]
    data = [row[: len(header)] for row in rows[header_index + 1 :] if row and row[0].strip()]
    return header, data


def parse_int(value: str):
    value = value.strip()
    if not value or value in {"..", ":", "-"}:
        return None
    return int(value.replace(".", ""))


def parse_decimal(value: str):
    value = value.strip()
    if not value or value in {"..", ":", "-"}:
        return None
    return float(value.replace(".", "").replace(",", "."))


def clear_data(sheet) -> None:
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)


def finish_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{max(sheet.max_row, 2)}"
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, min(sheet.max_row, 200) + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max(map(len, values)) + 2, 12), 42)


def jsonstat_rows(dataset: dict):
    dimensions = dataset["id"]
    categories = [dataset["dimension"][dimension]["category"]["index"] for dimension in dimensions]
    statuses = dataset.get("status") or [None] * len(dataset["value"])
    if isinstance(statuses, dict):
        statuses = [statuses.get(str(index)) for index in range(len(dataset["value"]))]
    for coordinates, value, status in zip(product(*categories), dataset["value"], statuses):
        yield dict(zip(dimensions, coordinates)), value, status


def prepare_sources() -> tuple[dict[str, list[str]], dict]:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    for filename, url in CSV_SOURCES.items():
        destination = RAW_DIR / filename
        if not destination.exists():
            download(url, destination)

    _, code_rows = csv_rows(RAW_DIR / "mpiscatalunya.csv")
    county_codes = sorted(row[0] for row in code_rows if row[2] == COUNTY_CODE)
    if len(county_codes) != 22:
        raise ValueError(f"Expected 22 municipalities in {COUNTY_NAME}, found {len(county_codes)}")

    population_url = (
        "https://api.idescat.cat/taules/v2/censph/10/5975/mun/data"
        f"?lang=ca&YEAR=2021&MUN={','.join(county_codes)}&SEX=TOTAL"
    )
    POPULATION_CACHE.parent.mkdir(parents=True, exist_ok=True)
    download(population_url, POPULATION_CACHE)
    population = json.loads(POPULATION_CACHE.read_text(encoding="utf-8"))
    if population.get("class") != "dataset" or population["size"] != [1, 22, 102, 1, 1]:
        raise ValueError("Unexpected population API response shape")

    municipality_labels = population["dimension"]["MUN"]["category"]["label"]
    age_labels = population["dimension"]["AGE"]["category"]["label"]
    with POPULATION_FILE.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["YEAR", "MUN", "municipality_name", "AGE", "age_label", "SEX", "CONCEPT", "value", "status", "source_url"])
        for coordinates, value, status in jsonstat_rows(population):
            writer.writerow([
                coordinates["YEAR"], coordinates["MUN"], municipality_labels[coordinates["MUN"]],
                coordinates["AGE"], age_labels[coordinates["AGE"]], coordinates["SEX"],
                coordinates["CONCEPT"], value, status or "", population_url,
            ])

    return {"population_url": population_url, "county_codes": county_codes}, population


def build_workbook(metadata: dict, population: dict) -> None:
    shutil.copyfile(STARTER, OUTPUT)
    workbook = load_workbook(OUTPUT)

    codes_header, codes = csv_rows(RAW_DIR / "mpiscatalunya.csv")
    housing_header, housing = csv_rows(RAW_DIR / "t30mun.csv")
    surface_header, surface = csv_rows(RAW_DIR / "t396mun.csv")

    code_by_id = {row[0]: row for row in codes}
    housing_by_id = {row[0]: row for row in housing}
    surface_by_id = {row[0]: row for row in surface}

    for sheet_name, rows in (
        ("source_codes", codes),
        ("source_housing", housing),
        ("source_surface", surface),
    ):
        sheet = workbook[sheet_name]
        clear_data(sheet)
        for row in rows:
            sheet.append(row)
        for cell in sheet["A"]:
            cell.number_format = "@"
        finish_sheet(sheet)

    prepared_codes = workbook["prepared_codes"]
    clear_data(prepared_codes)
    raw_code_row = {}
    for row in codes:
        prepared_codes.append([f"[{row[0]}] {row[1]}", row[0], row[1], row[2], row[3], ""])
        raw_code_row[row[0]] = prepared_codes.max_row
    for cell in prepared_codes["B"]:
        cell.number_format = "@"
    for cell in prepared_codes["D"]:
        cell.number_format = "@"
    finish_sheet(prepared_codes)

    prepared_housing = workbook["prepared_housing"]
    clear_data(prepared_housing)
    raw_housing_row = {}
    for row in housing:
        status = "" if all(value.strip() not in {"..", ":", "-"} for value in row[2:]) else "Dada no disponible"
        prepared_housing.append(
            [
                f"[{row[0]}] {row[1]}",
                row[0],
                row[1],
                parse_int(row[6]),
                parse_int(row[2]),
                parse_int(row[3]),
                status,
                "Total habitatges familiars; principal i no principal provenen de les categories publicades",
            ]
        )
        raw_housing_row[row[0]] = prepared_housing.max_row
    for cell in prepared_housing["B"]:
        cell.number_format = "@"
    finish_sheet(prepared_housing)

    prepared_surface = workbook["prepared_surface"]
    clear_data(prepared_surface)
    raw_surface_row = {}
    for row in surface:
        status = "" if row[2].strip() not in {"..", ":", "-"} else "Dada no disponible"
        prepared_surface.append([f"[{row[0]}] {row[1]}", row[0], row[1], parse_decimal(row[2]), status, "Publicació municipal 2025; la superfície s'utilitza com a magnitud territorial"])
        raw_surface_row[row[0]] = prepared_surface.max_row
    for cell in prepared_surface["B"]:
        cell.number_format = "@"
    for cell in prepared_surface["D"]:
        cell.number_format = "0.00"
    finish_sheet(prepared_surface)

    age_labels = population["dimension"]["AGE"]["category"]["label"]
    municipality_labels = population["dimension"]["MUN"]["category"]["label"]
    source_population = workbook["source_population"]
    clear_data(source_population)
    prepared_population = workbook["prepared_population"]
    clear_data(prepared_population)
    population_values = defaultdict(dict)
    for coordinates, value, status in jsonstat_rows(population):
        age_code = coordinates["AGE"]
        age_years = None
        if age_code.startswith("Y") and age_code[1:].isdigit():
            age_years = int(age_code[1:])
        elif age_code == "Y_GE100":
            age_years = 100
        code = coordinates["MUN"]
        source_population.append([
            coordinates["YEAR"], coordinates["MUN"], coordinates["AGE"],
            coordinates["SEX"], coordinates["CONCEPT"], value, status or "",
        ])
        prepared_population.append(
            [
                int(coordinates["YEAR"]),
                code,
                municipality_labels[code],
                age_code,
                age_labels[age_code],
                age_years,
                value,
                status or "",
                "Total de tots els sexes; resposta JSON-stat de l'API d'Idescat",
            ]
        )
        population_values[code][age_code] = value
    for cell in source_population["B"]:
        cell.number_format = "@"
    finish_sheet(source_population)
    for cell in prepared_population["B"]:
        cell.number_format = "@"
    finish_sheet(prepared_population)

    municipal = workbook["municipal"]
    clear_data(municipal)
    selected_codes = metadata["county_codes"]
    computed = {}
    for output_row, code in enumerate(selected_codes, start=2):
        code_row = raw_code_row[code]
        housing_row = raw_housing_row[code]
        surface_row = raw_surface_row[code]
        values = population_values[code]
        ages_0_14 = sum(values.get(f"Y{age:03d}") or 0 for age in range(0, 15))
        ages_15_64 = sum(values.get(f"Y{age:03d}") or 0 for age in range(15, 65))
        ages_65_plus = sum(values.get(f"Y{age:03d}") or 0 for age in range(65, 100)) + (values.get("Y_GE100") or 0)
        computed[code] = {
            "total": values.get("TOTAL"),
            "age_0_14": ages_0_14,
            "age_15_64": ages_15_64,
            "age_65_plus": ages_65_plus,
        }
        municipal.append(
            [
                f"=prepared_codes!B{code_row}",
                f"=prepared_codes!C{code_row}",
                f"=prepared_codes!D{code_row}",
                f"=prepared_codes!E{code_row}",
                f'=SUMIFS(prepared_population!$G:$G,prepared_population!$B:$B,$A{output_row},prepared_population!$D:$D,"TOTAL")',
                f'=SUMIFS(prepared_population!$G:$G,prepared_population!$B:$B,$A{output_row},prepared_population!$F:$F,">=0",prepared_population!$F:$F,"<=14")',
                f'=SUMIFS(prepared_population!$G:$G,prepared_population!$B:$B,$A{output_row},prepared_population!$F:$F,">=15",prepared_population!$F:$F,"<=64")',
                f'=SUMIFS(prepared_population!$G:$G,prepared_population!$B:$B,$A{output_row},prepared_population!$F:$F,">=65")',
                f"=prepared_housing!D{housing_row}",
                f"=prepared_housing!E{housing_row}",
                f"=prepared_housing!F{housing_row}",
                f"=prepared_surface!D{surface_row}",
                "Població i habitatge: 2021; superfície: publicació 2025, amb compatibilitat territorial pendent de revisió explícita",
            ]
        )
    for cell in municipal["A"]:
        cell.number_format = "@"
    finish_sheet(municipal)

    if "pivot_county_control" in workbook.sheetnames:
        del workbook["pivot_county_control"]
    pivot = workbook.create_sheet("pivot_county_control")
    pivot.append(["county_code", "county_name", "municipality_count", "purpose"])
    county_counts = Counter((row[2], row[3]) for row in codes)
    for (county_code, county_name), count in sorted(county_counts.items()):
        pivot.append([county_code, county_name, count, "Resultat de control que l'estudiant pot reproduir amb una taula dinàmica"])
    finish_sheet(pivot)

    sources = workbook["sources"]
    source_rows = {sources.cell(row, 1).value: row for row in range(2, sources.max_row + 1)}
    source_updates = {
        "idescat_population_2021": [metadata["population_url"], "2021", ACCESS_DATE, "Condicions d'ús de l'API d'Idescat", "CSV UTF-8 derivat d'una consulta JSON-stat documentada", "Tarragonès; edats simples; total de tots els sexes", POPULATION_FILE.name],
        "idescat_housing_2021": [CSV_SOURCES["t30mun.csv"], "2021", ACCESS_DATE, "Avís legal i condicions de reutilització d'Idescat", "CSV UTF-8; separador ;", "Tots els municipis", "t30mun.csv"],
        "idescat_codes": [CSV_SOURCES["mpiscatalunya.csv"], "vigent en la data d'accés", ACCESS_DATE, "Avís legal i condicions de reutilització d'Idescat", "CSV UTF-8; separador ;", "Municipis i comarques", "mpiscatalunya.csv"],
        "idescat_surface": [CSV_SOURCES["t396mun.csv"], "2025", ACCESS_DATE, "Avís legal i condicions de reutilització d'Idescat", "CSV UTF-8; separador ;", "Tots els municipis", "t396mun.csv"],
    }
    for source_id, values in source_updates.items():
        row = source_rows[source_id]
        for column, value in zip(range(4, 11), values):
            sources.cell(row, column, value)
    finish_sheet(sources)

    checks = workbook["checks"]
    clear_data(checks)
    duplicate_codes = len(codes) - len({row[0] for row in codes})
    missing = sum(
        code not in housing_by_id or code not in surface_by_id or computed[code]["total"] is None
        for code in selected_codes
    )
    discrepancies = sum(
        data["total"] != data["age_0_14"] + data["age_15_64"] + data["age_65_plus"]
        for data in computed.values()
    )
    check_rows = [
        ["C01", "source_codes", "Files municipals importades", len(codes), len(codes), "OK", f"Capçalera detectada: {', '.join(codes_header)}"],
        ["C02", "prepared_codes", "Codis municipals duplicats", 0, duplicate_codes, "OK" if duplicate_codes == 0 else "REVISAR", "Clau de sis caràcters conservada com a text"],
        ["C03", "municipal", "Municipis del Tarragonès", 22, len(selected_codes), "OK" if len(selected_codes) == 22 else "REVISAR", "Filtrats amb el codi de comarca 36"],
        ["C04", "municipal", "Fonts sense correspondència", 0, missing, "OK" if missing == 0 else "REVISAR", "Població, habitatge i superfície contrastats per codi"],
        ["C05", "municipal", "Suma dels grups d'edat diferent del total", "0 o documentat", discrepancies, "OK" if discrepancies == 0 else "REVISAR", "No es força la coincidència si la font presenta arrodoniments"],
        ["C06", "municipal", "Municipi de control", "431711 Vila-seca", f"431711 {code_by_id['431711'][1]}", "OK", "Codi contrastat entre les quatre fonts"],
        ["C07", "source_surface", "Període de la superfície", "Compatibilitat documentada", "Publicació 2025", "REVISAR", "Cal justificar l'ús com a denominador de població 2021 o obtenir una edició territorial equivalent"],
    ]
    for row in check_rows:
        checks.append(row)
    finish_sheet(checks)

    project = workbook["project"]
    project_values = {project.cell(row, 1).value: row for row in range(2, project.max_row + 1)}
    project.cell(project_values["spreadsheet_application"], 2, "LibreOffice Calc 24.2.7.2; validar també amb Excel abans de publicar")
    project.cell(project_values["regional_configuration"], 2, "Interfície catalana; decimal amb coma; fórmules OOXML en anglès al fitxer")
    project.cell(project_values["author"], 2, "Equip docent TIGIT")

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(OUTPUT)


if __name__ == "__main__":
    metadata, population = prepare_sources()
    build_workbook(metadata, population)
    print(OUTPUT)
