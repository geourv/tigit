#!/usr/bin/env python3
"""Publish visible student and teaching snapshots under dist/."""

from __future__ import annotations

import csv
import shutil
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PROCESSED = PROJECT_ROOT / "data" / "processed"
DIST = PROJECT_ROOT / "dist"

SNAPSHOTS = (
    (
        "student", "chapter-01", PROCESSED / "tigit-01-preparacio-dades.xlsx",
        "tigit-01-preparacio-dades.xlsx",
    ),
    (
        "teaching", "chapter-01", PROCESSED / "tigit-01-preparacio-dades-teaching.xlsx",
        "tigit-01-preparacio-dades.xlsx",
    ),
    (
        "teaching", "chapter-02", PROCESSED / "tigit-02-indicadors-territorials-teaching.xlsx",
        "tigit-02-indicadors-territorials.xlsx",
    ),
    (
        "teaching", "chapter-03", PROCESSED / "tigit-03-semiologia-visualitzacio-teaching.xlsx",
        "tigit-03-semiologia-visualitzacio.xlsx",
    ),
    (
        "teaching", "chapter-07", PROCESSED / "tigit-07-teoria-color-teaching.xlsx",
        "tigit-07-teoria-color.xlsx",
    ),
)


def main() -> None:
    manifest = []
    previous_teaching_sheets: set[str] = set()
    for role, chapter, source, filename in SNAPSHOTS:
        destination = DIST / role / chapter / filename
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(source, destination)
        sheets = load_workbook(source, read_only=True).sheetnames
        added = sorted(set(sheets) - previous_teaching_sheets) if role == "teaching" else sheets
        if role == "teaching":
            previous_teaching_sheets = set(sheets)
        manifest.append({
            "role": role,
            "chapter": chapter,
            "file": destination.relative_to(DIST).as_posix(),
            "sheet_count": len(sheets),
            "sheets_added_from_previous_teaching_snapshot": ",".join(added),
        })

    with (DIST / "snapshot-manifest.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=manifest[0].keys())
        writer.writeheader()
        writer.writerows(manifest)
    print(DIST / "snapshot-manifest.csv")


if __name__ == "__main__":
    main()
