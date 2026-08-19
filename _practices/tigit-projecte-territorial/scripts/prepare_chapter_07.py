#!/usr/bin/env python3
"""Create the chapter 7 color-registry workbook and recolor charts."""

from __future__ import annotations

import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REPOSITORY_ROOT = PROJECT_ROOT.parents[1]
SOURCE = PROJECT_ROOT / "data" / "processed" / "tigit-03-semiologia-visualitzacio-teaching.xlsx"
OUTPUT = PROJECT_ROOT / "data" / "processed" / "tigit-07-teoria-color-teaching.xlsx"
TMP_DIR = REPOSITORY_ROOT / "tmp" / "chapter-07"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

PALETTE_ROWS = [
    ("context", "project", "background", "Fons", "#FFFFFF", "Disseny del projecte", "Pàgina i gràfics"),
    ("context", "project", "main_text", "Text principal", "#222222", "Disseny del projecte", "Títols i etiquetes"),
    ("context", "project", "secondary_text", "Text secundari", "#595959", "Disseny del projecte", "Fonts i notes"),
    ("context", "project", "grid", "Retícula", "#D9D9D9", "Disseny del projecte", "Guies i separadors"),
    ("context", "project", "neutral_data", "Dada neutral", "#7A8C99", "Disseny del projecte", "Municipis comparables"),
    ("context", "project", "vila_seca_accent", "Accent Vila-seca", "#D55E00", "Disseny del projecte", "Accent redundant amb etiqueta"),
    ("context", "project", "no_data", "Sense dades", "#BDBDBD", "Disseny del projecte", "Nuls, amb etiqueta o patró"),
    ("age", "ColorBrewer Set2", "young", "0–14", "#66C2A5", "ColorBrewer", "Estructura d'edats"),
    ("age", "ColorBrewer Set2", "working_age", "15–64", "#FC8D62", "ColorBrewer", "Estructura d'edats"),
    ("age", "ColorBrewer Set2", "older", "65+", "#8DA0CB", "ColorBrewer", "Estructura d'edats"),
    ("map_bugn", "ColorBrewer BuGn 5", "class_1", "Molt baix", "#EDF8FB", "ColorBrewer", "Candidata seqüencial"),
    ("map_bugn", "ColorBrewer BuGn 5", "class_2", "Baix", "#B2E2E2", "ColorBrewer", "Candidata seqüencial"),
    ("map_bugn", "ColorBrewer BuGn 5", "class_3", "Intermedi", "#66C2A4", "ColorBrewer", "Candidata seqüencial"),
    ("map_bugn", "ColorBrewer BuGn 5", "class_4", "Alt", "#2CA25F", "ColorBrewer", "Candidata seqüencial"),
    ("map_bugn", "ColorBrewer BuGn 5", "class_5", "Molt alt", "#006D2C", "ColorBrewer", "Candidata seqüencial"),
    ("map_rdbu", "ColorBrewer RdBu 5", "negative_2", "Molt per sota", "#CA0020", "ColorBrewer", "Candidata divergent"),
    ("map_rdbu", "ColorBrewer RdBu 5", "negative_1", "Per sota", "#F4A582", "ColorBrewer", "Candidata divergent"),
    ("map_rdbu", "ColorBrewer RdBu 5", "reference", "Referència", "#F7F7F7", "ColorBrewer", "Candidata divergent"),
    ("map_rdbu", "ColorBrewer RdBu 5", "positive_1", "Per sobre", "#92C5DE", "ColorBrewer", "Candidata divergent"),
    ("map_rdbu", "ColorBrewer RdBu 5", "positive_2", "Molt per sobre", "#0571B0", "ColorBrewer", "Candidata divergent"),
]


def recalculated_values():
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    target = TMP_DIR / SOURCE.name
    if target.exists():
        target.unlink()
    subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", str(TMP_DIR), str(SOURCE)], check=True)
    return load_workbook(target, data_only=True)


def formula(row: int, channel_column: str) -> str:
    return (
        f'=IF({channel_column}{row}="","",IF({channel_column}{row}/255<=0.04045,'
        f'{channel_column}{row}/255/12.92,POWER(({channel_column}{row}/255+0.055)/1.055,2.4)))'
    )


def set_series_fill(series, color: str) -> None:
    series.graphicalProperties.solidFill = color
    series.graphicalProperties.line.solidFill = color


def build() -> None:
    calculated = recalculated_values()
    shutil.copyfile(SOURCE, OUTPUT)
    workbook = load_workbook(OUTPUT)
    project = workbook["project"]
    project_rows = {project.cell(row, 1).value: row for row in range(2, project.max_row + 1)}
    project.cell(project_rows["chapter_snapshot"], 2, "07")

    if "palette" in workbook.sheetnames:
        del workbook["palette"]
    palette = workbook.create_sheet("palette")
    palette.append([
        "palette_id", "family_or_tool", "role", "color_name", "HEX", "R", "G", "B",
        "linear_R", "linear_G", "linear_B", "relative_luminance", "contrast_white",
        "contrast_black", "origin", "intended_use", "grayscale_check", "cvd_check", "notes",
    ])
    for row_index, row in enumerate(PALETTE_ROWS, start=2):
        palette_id, family, role, name, hex_code, origin, intended_use = row
        palette.append([
            palette_id, family, role, name, hex_code,
            f'=IF(E{row_index}="","",HEX2DEC(MID(E{row_index},2,2)))',
            f'=IF(E{row_index}="","",HEX2DEC(MID(E{row_index},4,2)))',
            f'=IF(E{row_index}="","",HEX2DEC(MID(E{row_index},6,2)))',
            formula(row_index, "F"), formula(row_index, "G"), formula(row_index, "H"),
            f'=IF(I{row_index}="","",0.2126*I{row_index}+0.7152*J{row_index}+0.0722*K{row_index})',
            f'=IF(L{row_index}="","",1.05/(L{row_index}+0.05))',
            f'=IF(L{row_index}="","",(L{row_index}+0.05)/0.05)',
            origin, intended_use, "PENDING", "PENDING", "",
        ])
        palette.cell(row_index, 5).fill = PatternFill("solid", fgColor=hex_code.lstrip("#"))

    for slot in range(1, 6):
        row_index = palette.max_row + 1
        palette.append([
            "adobe_candidate", "Adobe Color", f"candidate_{slot}", f"Candidat {slot}", "",
            f'=IF(E{row_index}="","",HEX2DEC(MID(E{row_index},2,2)))',
            f'=IF(E{row_index}="","",HEX2DEC(MID(E{row_index},4,2)))',
            f'=IF(E{row_index}="","",HEX2DEC(MID(E{row_index},6,2)))',
            formula(row_index, "F"), formula(row_index, "G"), formula(row_index, "H"),
            f'=IF(I{row_index}="","",0.2126*I{row_index}+0.7152*J{row_index}+0.0722*K{row_index})',
            f'=IF(L{row_index}="","",1.05/(L{row_index}+0.05))',
            f'=IF(L{row_index}="","",(L{row_index}+0.05)/0.05)',
            "https://color.adobe.com/", "Context o accent; no assignar directament a classes", "PENDING", "PENDING", "Enganxar un codi HEX explorat",
        ])

    for cell in palette[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True)
    palette.freeze_panes = "A2"
    palette.auto_filter.ref = palette.dimensions
    for column in range(1, palette.max_column + 1):
        palette.column_dimensions[get_column_letter(column)].width = 18
    for column in (12, 13, 14):
        for cell in palette[get_column_letter(column)][1:]:
            cell.number_format = "0.00"

    age_chart = workbook["chart_01_age_structure"]._charts[0]
    for series, color in zip(age_chart.series, ("66C2A5", "FC8D62", "8DA0CB")):
        set_series_fill(series, color)

    housing_chart = workbook["chart_02_nonprincipal"]._charts[0]
    set_series_fill(housing_chart.series[0], "7A8C99")
    values = calculated["charts_data"]
    vila_seca_index = next(row - 2 for row in range(2, 24) if values.cell(row, 7).value == "431711")
    accent = DataPoint(idx=vila_seca_index)
    accent.graphicalProperties.solidFill = "D55E00"
    accent.graphicalProperties.line.solidFill = "D55E00"
    housing_chart.series[0].dPt = [accent]

    scatter = workbook["chart_03_scatter"]._charts[0].series[0]
    scatter.marker.graphicalProperties.solidFill = "006D2C"
    scatter.marker.graphicalProperties.line.solidFill = "00441B"

    pyramid = workbook["chart_06_population_pyramid"]._charts[0]
    for series, color in zip(pyramid.series, ("0571B0", "CA0020")):
        set_series_fill(series, color)

    histogram = workbook["chart_07_housing_histogram"]._charts[0]
    set_series_fill(histogram.series[0], "66C2A4")

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    build()
