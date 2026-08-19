#!/usr/bin/env python3
"""Create the chapter 3 workbook with editable linked charts."""

from __future__ import annotations

import csv
import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference, ScatterChart, Series
from openpyxl.styles import Font


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REPOSITORY_ROOT = PROJECT_ROOT.parents[1]
SOURCE = PROJECT_ROOT / "data" / "processed" / "tigit-02-indicadors-territorials-teaching.xlsx"
OUTPUT = PROJECT_ROOT / "data" / "processed" / "tigit-03-semiologia-visualitzacio-teaching.xlsx"
TMP_DIR = REPOSITORY_ROOT / "tmp" / "chapter-03"
POPULATION_TIME_CSV = PROJECT_ROOT / "data" / "raw" / "poblacio-vila-seca-2000-2022.csv"
POPULATION_PYRAMID_CSV = PROJECT_ROOT / "data" / "raw" / "poblacio-vila-seca-edat-sexe-2021.csv"


def recalculated_values():
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    recalculated = TMP_DIR / SOURCE.name
    if recalculated.exists():
        recalculated.unlink()
    subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", str(TMP_DIR), str(SOURCE)], check=True)
    return load_workbook(recalculated, data_only=True)


def replace_sheet(workbook, title):
    if title in workbook.sheetnames:
        del workbook[title]
    return workbook.create_sheet(title)


def add_age_chart(sheet, data_sheet, anchor="A1"):
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "percentStacked"
    chart.overlap = 100
    chart.title = "Estructura de la població per grans grups d'edat"
    chart.x_axis.title = "% de població"
    chart.y_axis.title = "Municipi"
    chart.height = 18
    chart.width = 28
    chart.add_data(Reference(data_sheet, min_col=3, max_col=5, min_row=1, max_row=23), titles_from_data=True)
    chart.set_categories(Reference(data_sheet, min_col=2, min_row=2, max_row=23))
    sheet.add_chart(chart, anchor)


def add_housing_chart(sheet, data_sheet, anchor="A1", horizontal=True):
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.title = "Pes de l'habitatge no principal als municipis del Tarragonès"
    chart.x_axis.title = "% sobre el total d'habitatges" if horizontal else "Municipi"
    chart.y_axis.title = "Municipi" if horizontal else "% sobre el total"
    chart.height = 18
    chart.width = 28
    chart.add_data(Reference(data_sheet, min_col=9, min_row=1, max_row=23), titles_from_data=True)
    chart.set_categories(Reference(data_sheet, min_col=8, min_row=2, max_row=23))
    sheet.add_chart(chart, anchor)


def build() -> None:
    values = recalculated_values()
    shutil.copyfile(SOURCE, OUTPUT)
    workbook = load_workbook(OUTPUT)
    project = workbook["project"]
    project_rows = {project.cell(row, 1).value: row for row in range(2, project.max_row + 1)}
    project.cell(project_rows["chapter_snapshot"], 2, "03")

    with POPULATION_TIME_CSV.open(encoding="utf-8", newline="") as handle:
        time_reader = csv.DictReader(handle)
        time_headers = time_reader.fieldnames
        time_records = list(time_reader)
    source_population_time = replace_sheet(workbook, "source_population_time")
    source_population_time.append(time_headers)
    for record in time_records:
        source_population_time.append([record.get(header) for header in time_headers])
    source_population_time.freeze_panes = "A2"

    prepared_population_time = replace_sheet(workbook, "prepared_population_time")
    prepared_population_time.append(["Any", "Població", "Canvi anual", "Canvi anual (%)", "Etiqueta biennal"])
    for output_row, record in enumerate(time_records, start=2):
        prepared_population_time.append([
            int(record["year"]), int(float(record["population"])),
            "" if output_row == 2 else f"=B{output_row}-B{output_row - 1}",
            "" if output_row == 2 else f'=IF(B{output_row - 1}>0,(B{output_row}/B{output_row - 1}-1)*100,NA())',
            f'=IF(OR(MOD(A{output_row},2)=0,A{output_row}=2022),A{output_row},"")',
        ])
    prepared_population_time.freeze_panes = "A2"

    with POPULATION_PYRAMID_CSV.open(encoding="utf-8", newline="") as handle:
        pyramid_reader = csv.DictReader(handle)
        pyramid_headers = pyramid_reader.fieldnames
        pyramid_records = list(pyramid_reader)
    source_population_pyramid = replace_sheet(workbook, "source_population_pyramid")
    source_population_pyramid.append(pyramid_headers)
    for record in pyramid_records:
        source_population_pyramid.append([record.get(header) for header in pyramid_headers])
    source_population_pyramid.freeze_panes = "A2"

    prepared_population_pyramid = replace_sheet(workbook, "prepared_population_pyramid")
    prepared_population_pyramid.append(["year", "age_years", "sex", "population"])
    for record in pyramid_records:
        prepared_population_pyramid.append([
            int(record["year"]), int(record["age_years"]), record["sex"], int(float(record["population"])),
        ])
    replace_sheet(workbook, "pivot_population_age_sex")

    pyramid_data = replace_sheet(workbook, "pyramid_data")
    pyramid_data.append(["Grup d'edat", "Edat inicial", "Edat final", "Homes", "Dones"])
    age_groups = [(f"{start}-{start + 4}", start, start + 4) for start in range(0, 85, 5)] + [("85+", 85, 120)]
    for output_row, (label, start, end) in enumerate(age_groups, start=2):
        pyramid_data.append([
            label, start, end,
            f'=-SUMIFS(prepared_population_pyramid!$D:$D,prepared_population_pyramid!$C:$C,"M",prepared_population_pyramid!$B:$B,">="&B{output_row},prepared_population_pyramid!$B:$B,"<="&C{output_row})',
            f'=SUMIFS(prepared_population_pyramid!$D:$D,prepared_population_pyramid!$C:$C,"F",prepared_population_pyramid!$B:$B,">="&B{output_row},prepared_population_pyramid!$B:$B,"<="&C{output_row})',
        ])

    histogram_data = replace_sheet(workbook, "histogram_data")
    histogram_data.append(["Límit inferior", "Límit superior", "Interval (%)", "Municipis"])
    for output_row, lower in enumerate(range(0, 70, 10), start=2):
        upper = lower + 10
        histogram_data.append([
            lower, upper, f"{lower}-{upper}",
            f'=COUNTIFS(indicators_housing!$G$2:$G$23,">="&A{output_row},indicators_housing!$G$2:$G$23,"<"&B{output_row})',
        ])

    data = replace_sheet(workbook, "charts_data")
    data.append(["municipality_code", "municipality", "age_0_14_pct", "age_15_64_pct", "age_65_plus_pct", "age_sum_check", "housing_code", "housing_municipality", "housing_non_main_pct", "scatter_code", "scatter_municipality", "age_65_plus_pct_x", "housing_non_main_pct_y"])

    demographic = values["indicators_demography"]
    housing = values["indicators_housing"]
    source_rows = list(range(2, 24))
    age_order = sorted(source_rows, key=lambda row: demographic.cell(row, 8).value or 0)
    housing_order = sorted(source_rows, key=lambda row: housing.cell(row, 7).value or 0, reverse=True)
    for output_row in range(2, 24):
        age_row = age_order[output_row - 2]
        housing_row = housing_order[output_row - 2]
        data.append([
            f"=indicators_demography!A{age_row}", f"=indicators_demography!B{age_row}",
            f"=indicators_demography!G{age_row}",
            f"=municipal!G{age_row}/municipal!E{age_row}*100",
            f"=indicators_demography!H{age_row}", f"=SUM(C{output_row}:E{output_row})",
            f"=indicators_housing!A{housing_row}", f"=indicators_housing!B{housing_row}",
            f"=indicators_housing!G{housing_row}",
            f"=indicators_housing!A{output_row}", f"=indicators_housing!B{output_row}",
            f"=indicators_demography!H{output_row}", f"=indicators_housing!G{output_row}",
        ])
    data["O1"] = "housing_type"
    data["P1"] = "housing_count"
    data["O2"] = "Habitatges principals"
    data["P2"] = "=SUM(indicators_housing!E2:E23)"
    data["O3"] = "Habitatges no principals"
    data["P3"] = "=SUM(indicators_housing!F2:F23)"
    data.freeze_panes = "A2"
    data.auto_filter.ref = "A1:M23"

    audit = replace_sheet(workbook, "chart_00_audit")
    audit["A1"] = "Versió inicial"
    audit["A1"].font = Font(bold=True)
    audit["A32"] = "Versió revisada"
    audit["A32"].font = Font(bold=True)
    add_housing_chart(audit, data, "A3", horizontal=False)
    add_housing_chart(audit, data, "A34", horizontal=True)

    age_sheet = replace_sheet(workbook, "chart_01_age_structure")
    add_age_chart(age_sheet, data)
    housing_sheet = replace_sheet(workbook, "chart_02_nonprincipal")
    add_housing_chart(housing_sheet, data)

    scatter_sheet = replace_sheet(workbook, "chart_03_scatter")
    scatter = ScatterChart()
    scatter.title = "Envelliment i habitatge no principal"
    scatter.x_axis.title = "Població de 65 anys o més (%)"
    scatter.y_axis.title = "Habitatge no principal (%)"
    scatter.height = 17
    scatter.width = 25
    xvalues = Reference(data, min_col=12, min_row=2, max_row=23)
    yvalues = Reference(data, min_col=13, min_row=2, max_row=23)
    scatter_series = Series(yvalues, xvalues, title="Municipis")
    scatter_series.marker.symbol = "circle"
    scatter_series.marker.size = 7
    scatter_series.marker.graphicalProperties.solidFill = "4472C4"
    scatter_series.marker.graphicalProperties.line.solidFill = "2F5597"
    scatter_series.graphicalProperties.line.noFill = True
    scatter.series.append(scatter_series)
    scatter_sheet.add_chart(scatter, "A1")

    donut_sheet = replace_sheet(workbook, "chart_04_housing_donut")
    donut = DoughnutChart()
    donut.title = "Composició del parc d'habitatges del Tarragonès"
    donut.holeSize = 55
    donut.height = 15
    donut.width = 20
    donut.add_data(Reference(data, min_col=16, min_row=1, max_row=3), titles_from_data=True)
    donut.set_categories(Reference(data, min_col=15, min_row=2, max_row=3))
    donut_sheet.add_chart(donut, "A1")

    line_sheet = replace_sheet(workbook, "chart_05_population_time")
    line = LineChart()
    line.title = "Evolució de la població de Vila-seca"
    line.x_axis.title = "Any"
    line.y_axis.title = "Població"
    line.height = 16
    line.width = 28
    line.add_data(Reference(prepared_population_time, min_col=2, min_row=1, max_row=prepared_population_time.max_row), titles_from_data=True)
    line.set_categories(Reference(prepared_population_time, min_col=5, min_row=2, max_row=prepared_population_time.max_row))
    for series, color in zip(line.series, ("4472C4",)):
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 28575
        series.marker.symbol = "circle"
        series.marker.size = 6
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color
    line_sheet.add_chart(line, "A1")

    pyramid_sheet = replace_sheet(workbook, "chart_06_population_pyramid")
    pyramid = BarChart()
    pyramid.type = "bar"
    pyramid.grouping = "stacked"
    pyramid.overlap = 100
    pyramid.title = "Piràmide de població de Vila-seca, 2021"
    pyramid.x_axis.title = "Grup d'edat"
    pyramid.y_axis.title = "Població"
    pyramid.height = 18
    pyramid.width = 25
    pyramid.add_data(Reference(pyramid_data, min_col=4, max_col=5, min_row=1, max_row=pyramid_data.max_row), titles_from_data=True)
    pyramid.set_categories(Reference(pyramid_data, min_col=1, min_row=2, max_row=pyramid_data.max_row))
    pyramid_sheet.add_chart(pyramid, "A1")

    histogram_sheet = replace_sheet(workbook, "chart_07_housing_histogram")
    histogram = BarChart()
    histogram.type = "col"
    histogram.gapWidth = 0
    histogram.title = "Distribució municipal de l'habitatge no principal"
    histogram.x_axis.title = "Percentatge d'habitatge no principal"
    histogram.y_axis.title = "Nombre de municipis"
    histogram.legend = None
    histogram.height = 16
    histogram.width = 23
    histogram.add_data(Reference(histogram_data, min_col=4, min_row=1, max_row=histogram_data.max_row), titles_from_data=True)
    histogram.set_categories(Reference(histogram_data, min_col=3, min_row=2, max_row=histogram_data.max_row))
    histogram_sheet.add_chart(histogram, "A1")

    for sheet in (audit, age_sheet, housing_sheet, scatter_sheet, donut_sheet, line_sheet, pyramid_sheet, histogram_sheet):
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    build()
