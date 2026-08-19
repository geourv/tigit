#!/usr/bin/env python3
"""Check chapter 1 source files, workbook, and packages."""

from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
RAW_DIR = PROJECT_ROOT / "data" / "raw"
WORKBOOK = PROJECT_ROOT / "data" / "processed" / "tigit-01-preparacio-dades-teaching.xlsx"


def main() -> None:
    required_raw = {
        "mpiscatalunya.csv",
        "t30mun.csv",
        "t396mun.csv",
        "poblacio-edat-tarragones-2021.csv",
    }
    missing = sorted(name for name in required_raw if not (RAW_DIR / name).is_file())
    assert not missing, f"Missing raw files: {missing}"

    import csv
    with (RAW_DIR / "poblacio-edat-tarragones-2021.csv").open(encoding="utf-8", newline="") as handle:
        population = list(csv.DictReader(handle))
    assert len(population) == 2244
    assert {row["YEAR"] for row in population} == {"2021"}
    assert {row["SEX"] for row in population} == {"TOTAL"}
    assert len({row["MUN"] for row in population}) == 22

    workbook = load_workbook(WORKBOOK, data_only=False)
    expected_sheets = {
        "project",
        "sources",
        "dictionary",
        "source_codes",
        "source_surface",
        "source_population",
        "source_housing",
        "prepared_codes",
        "prepared_surface",
        "prepared_population",
        "prepared_housing",
        "checks",
        "municipal",
        "pivot_county_control",
    }
    assert expected_sheets.issubset(workbook.sheetnames)
    assert workbook["municipal"].max_row == 23
    assert workbook["source_population"].max_row == 2245
    assert workbook["prepared_population"].max_row == 2245
    assert workbook["municipal"]["E2"].value.startswith("=SUMIFS(")
    vila_seca_row = next(
        row for row in range(2, workbook["prepared_codes"].max_row + 1)
        if workbook["prepared_codes"].cell(row, 2).value == "431711"
    )
    assert any(
        workbook["municipal"].cell(row, 1).value == f"=prepared_codes!B{vila_seca_row}"
        and workbook["municipal"].cell(row, 2).value == f"=prepared_codes!C{vila_seca_row}"
        for row in range(2, 24)
    )

    packages = (
        REPOSITORY_ROOT / "assets" / "downloads" / "tigit-01-preparacio-dades-student.zip",
        PROJECT_ROOT / "dist" / "course-packages" / "tigit-practiques-teaching.zip",
    )
    for package in packages:
        if package.exists():
            with ZipFile(package) as archive:
                assert archive.testzip() is None

    print("Chapter 1 checks passed")


if __name__ == "__main__":
    main()
